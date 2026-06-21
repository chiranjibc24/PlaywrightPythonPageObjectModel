from utils.jsonHandling import jsonData
import json
import os
import csv

from dotenv import load_dotenv
from openpyxl import load_workbook
import pytest


def test_cli_env():
    load_dotenv(os.getenv("file"))
    print(os.getenv("us1"))
    print(os.getenv("pw1"))


def test_json():
    with open("testData\\data1.json") as data:
         formattedData = json.load(data)
         print(formattedData["username"])


def test_json_1():
    testData = jsonData("testData\\data1.json")
    print(testData)

# @pytest.mark.datahandling
def test_csv():
    with open("testData\\credentails.csv") as data:
        formattedData  = csv.DictReader(data)
        values = []
        for i in formattedData:
            values.append(i)

    print(values[1]["username"])

# @pytest.mark.datahanding
def test_read_csv():
    with open("testData\\credentails.csv") as data:
        formattedData = csv.DictReader(data)
        values = list(formattedData)

    print(values[0]["username"])
        
#pip in install
# @pytest.mark.datahandling
def test_read_excel():
    workbook = load_workbook("testData/sample_creds.xlsx")
    sheet  = workbook["Sheet2"]
    values = []
    for i in sheet.iter_rows(min_col=1, values_only=2):
        values.append(i)
    print(values)

# @pytest.mark.datahandling
def test_write_excel():
    workbook = load_workbook("testData/sample_creds.xlsx")
    sheet = workbook.create_sheet("sheet5")
    # sheet  = workbook["sheet1"]
    # # sheet.append(["username123","password123"])
    sheet["A2"] = "12345"
    # sheet.delete_rows(3,sheet.max_row - 3)
    workbook.save("testData/sample_creds.xlsx")

@pytest.mark.datahandling
@pytest.mark.parametrize("a,b,expected",[(2,3,5),(4,5,9),(6,7,13)])  
def test_sum(a, b, expected):
    print(a),
    print(b)
    assert a + b == expected

